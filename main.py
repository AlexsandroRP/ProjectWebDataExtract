from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import *
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.service import Service as ChromeService
from time import sleep
from selenium.webdriver.support import expected_conditions
import openpyxl
import os
import pyautogui
import schedule
import pyperclip


def run_bot():
    def get_prices():
        def start_driver():
            chrome_options = Options()
            arguments = ['--lang=en-US', '--window-size=1300,1000',
                        '--incognito']

            for argument in arguments:
                chrome_options.add_argument(argument)

            chrome_options.add_experimental_option('prefs', {
                'download.prompt_for_download': False,
                'profile.default_content_setting_values.notifications': 2,
                'profile.default_content_setting_values.automatic_downloads': 1

            })

            driver = webdriver.Chrome(service=ChromeService(
                ChromeDriverManager().install()), options=chrome_options)
            wait = WebDriverWait(
                driver,
                10,
                poll_frequency=1,
                ignored_exceptions=[
                    NoSuchElementException,
                    ElementNotVisibleException,
                    ElementNotSelectableException,
                ]
            )
            return driver, wait
            
        driver, wait = start_driver()
        # First Site
        driver.get('https://www.kabum.com.br/busca/rtx-4080?page_number=1&page_size=20&facet_filters=&sort=price')   
        sleep(10)
        prices_one = wait.until(expected_conditions.visibility_of_all_elements_located((By.XPATH, "//span[@class='sc-3b515ca1-2 eqqhbT priceCard']")))
        price_gpu_1 = float(prices_one[0].text.split(' ')[1].replace('.','').replace(',','.'))

        # Second site
        driver.get('https://www.pichau.com.br/search?q=rtx%204080&sort=price-asc&rgpu=7185')
        sleep(10)
        prices_two = wait.until(expected_conditions.visibility_of_all_elements_located((By. XPATH, "//div[@class='jss229']")))
        price_gpu_2 = float(prices_two[0].text.split(' ')[1].replace('.','').replace(',','.'))

        return price_gpu_1, price_gpu_2
    
#------------------------------------------------------------------

    def generate_sheet():
        # Putting the data into an excel workbook
        gpu_1, gpu_2 = get_prices()
        cost_gpu_example = 5000
        site_1 = 'www.kabum.com.br'
        site_2 = 'www.pichau.com.br'
        workbook = openpyxl.Workbook()
        del workbook['Sheet']
        workbook.create_sheet('profit_margin') # create the page
        sheet_profit_margin = workbook['profit_margin'] #select page
        sheet_profit_margin.append(['Site','Cost','Price','Profit']) # Create the row
        sheet_profit_margin.append([site_1, cost_gpu_example, gpu_1, gpu_1-cost_gpu_example]) # Adding data in first row

        sheet_profit_margin.append([site_2, cost_gpu_example, gpu_2, gpu_2-cost_gpu_example]) # Adding data in first row

        workbook.save('profit margin.xlsx')

        #------------------------------------------------------------------

        # Extracting the data from sheet
        data = openpyxl.load_workbook('profit margin.xlsx')
        # reading data from excel
        sheet_profit = data['profit_margin'] # << page to extract the data
        profit_margin = ''
        for row in sheet_profit.iter_rows(min_row=1): # read each row from the sheet
            profit_margin += f'{row[0].value},{row[1].value},{row[2].value},{row[3].value}{os.linesep}' # < site, cost, price, profit

        
        return profit_margin    

    def send_data_whatsapp():
# Sending the message on whatsapp
        profit_margin = generate_sheet()
        pyautogui.hotkey('win')
        sleep(1)
        pyautogui.write('whatsapp')
        sleep(1)
        pyautogui.hotkey('enter')
        sleep(3)
        button = pyautogui.locateCenterOnScreen('Capture.PNG')
        sleep(1)
        pyautogui.moveTo(button[0], button[1],duration=1)
        pyautogui.write('pan')
        sleep(1)
        pyautogui.move(0,130)
        sleep(1)
        pyautogui.leftClick()
        pyautogui.leftClick()
        sleep(1)
        pyperclip.copy(profit_margin)
        pyautogui.hotkey('ctrl', 'v')
        sleep(1)
        pyautogui.hotkey('enter')
    send_data_whatsapp()

schedule.every().day.at("19:00").do(run_bot)

while True:
    schedule.run_pending()
    sleep(1)    